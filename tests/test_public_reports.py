import pandas as pd
from openpyxl import load_workbook

from scripts.update_public_reports import (
    build_forward_exit_summary,
    build_forward_summary,
    build_strategy_policy_report,
    load_telegram_delivery_events,
    write_workbook,
)


def test_forward_summary_groups_trade_results():
    trades = pd.DataFrame(
        [
            {"entry_type": "Standard Limit", "timeframe": "M30", "strategy": "FVG", "outcome": "WIN", "net_profit_usd": 3.0},
            {"entry_type": "Standard Limit", "timeframe": "M30", "strategy": "FVG", "outcome": "LOSS", "net_profit_usd": -1.0},
        ]
    )

    summary = build_forward_summary(trades)

    assert summary.to_dict("records") == [
        {
            "entry_type": "Standard Limit",
            "timeframe": "M30",
            "strategy": "FVG",
            "trades": 2,
            "wins": 1,
            "losses": 1,
            "winrate_pct": 50.0,
            "gross_profit_usd": 3.0,
            "gross_loss_usd": 1.0,
            "profit_factor": 3.0,
            "avg_win_usd": 3.0,
            "avg_loss_usd": 1.0,
            "payoff_ratio": 3.0,
            "net_profit_usd": 2.0,
        }
    ]


def test_public_report_workbook_contains_expected_sheets(tmp_path):
    standard_limit = pd.DataFrame([{"timeframe": "M30", "strategy": "FVG", "winrate": 72.2}])
    trades = pd.DataFrame([{"position_id": 1, "outcome": "WIN", "net_profit_usd": 3.0}])
    summary = pd.DataFrame([{"entry_type": "Standard Limit", "trades": 1, "winrate_pct": 100.0}])
    output = tmp_path / "forward_test_report.xlsx"

    write_workbook(
        output,
        standard_limit,
        trades,
        summary,
        pd.DataFrame(),
        pd.DataFrame(),
        pd.DataFrame(),
        "2026-07-16 12:00:00 WIB",
    )

    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == [
        "Overview", "Standard Limit Real Tick", "Forward Trades", "Forward Summary",
        "Forward Exit Summary", "Strategy Policy", "Telegram Delivery",
    ]


def test_exit_summary_and_policy_report_expose_payoff_risk():
    trades = pd.DataFrame(
        [
            {"position_id": 1, "entry_type": "WatchZone", "timeframe": "M30", "strategy": "FVG", "mt5_exit_comment": "SMC Bot Soft TP", "net_profit_usd": -2.0},
            {"position_id": 2, "entry_type": "WatchZone", "timeframe": "M30", "strategy": "OB", "mt5_exit_comment": "[tp 4000]", "net_profit_usd": 3.0},
        ]
    )
    summary = build_forward_summary(trades.assign(outcome=["LOSS", "WIN"]))
    exit_summary = build_forward_exit_summary(trades)
    policy = build_strategy_policy_report(summary)

    assert set(exit_summary["exit_type"]) == {"SOFT_TP", "TAKE_PROFIT"}
    assert policy.loc[policy["strategy"] == "FVG", "deployment_status"].item() == "BLOCKED"
    assert policy.loc[policy["strategy"] == "OB", "deployment_status"].item() == "CANDIDATE"


def test_telegram_delivery_events_reads_valid_json_lines(tmp_path):
    path = tmp_path / "telegram_delivery_events.jsonl"
    path.write_text('{"sent_at_wib":"2026-07-16T12:00:00+07:00","channel":"text","delivered":true,"message":"entry"}\ninvalid\n', encoding="utf-8")

    events = load_telegram_delivery_events(path)

    assert events.to_dict("records") == [{"sent_at_wib": "2026-07-16T12:00:00+07:00", "channel": "text", "delivered": True, "message": "entry"}]
