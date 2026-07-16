"""Build public forward-test and real-tick report artifacts from local evidence.

The source files remain local because they can contain account-specific runtime
state. This script exports only trade evidence and aggregate metrics intended
for publication under reports/.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
REPORTS_DIR = BASE_DIR / "reports"
DEFAULT_POSITIONS = BASE_DIR / "scratch" / "bot_positions_ss_matched_full.csv"
DEFAULT_DEALS = BASE_DIR / "scratch" / "raw_deals_july.csv"
DEFAULT_STANDARD_LIMIT = BASE_DIR / "data" / "real_tick_standard_limit_may2026_all_tf.csv"
DEFAULT_TELEGRAM_EVENTS = BASE_DIR / "data" / "telegram_delivery_events.jsonl"


def _as_int(value) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _signal_ticket_map() -> dict[int, dict]:
    """Map known MT5 tickets to the original planned entry, SL, TP and confidence."""
    mapping: dict[int, dict] = {}
    for path in sorted((BASE_DIR / "data").glob("sent_signals*.json")):
        if "master_merged" in path.name:
            continue
        try:
            signals = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        for signal in signals.values():
            if not isinstance(signal, dict):
                continue
            for suffix, option_key in (("a", "0.5"), ("b", "0.618")):
                ticket = _as_int(signal.get(f"ticket_{suffix}"))
                if ticket is None:
                    continue
                features = signal.get(f"features_{option_key}") or {}
                mapping[ticket] = {
                    "planned_entry": signal.get(f"price_{option_key}", features.get("entry_price")),
                    "planned_sl": features.get("sl_price"),
                    "planned_tp": features.get("tp_price"),
                    "signal_probability": signal.get(f"probability_{option_key}"),
                    "signal_source": path.name,
                }
    return mapping


def _format_time(value) -> str:
    if pd.isna(value):
        return ""
    try:
        return pd.to_datetime(int(value), unit="s", utc=True).tz_convert("Asia/Jakarta").strftime("%Y-%m-%d %H:%M:%S WIB")
    except (TypeError, ValueError, OverflowError):
        return ""


def build_forward_trades(positions_path: Path, deals_path: Path) -> pd.DataFrame:
    positions = pd.read_csv(positions_path)
    closed_values = positions["is_closed"].astype(str).str.strip().str.lower()
    positions = positions[closed_values.isin({"true", "1", "yes"})].copy()
    deals = pd.read_csv(deals_path)
    ticket_map = _signal_ticket_map()
    records = []

    for _, position in positions.iterrows():
        position_id = _as_int(position.get("position_id"))
        if position_id is None:
            continue
        position_deals = deals[deals["position_id"].map(_as_int) == position_id].sort_values("time_msc")
        entries = position_deals[position_deals["entry"] == 0]
        exits = position_deals[position_deals["entry"] != 0]
        entry = entries.iloc[0] if not entries.empty else None
        exit_deal = exits.iloc[-1] if not exits.empty else None

        signal = ticket_map.get(position_id, {})
        if entry is not None:
            signal = ticket_map.get(_as_int(entry.get("order")), signal)

        net_profit = float(position.get("net_profit", 0.0) or 0.0)
        is_win = bool(position.get("win", net_profit > 0))
        records.append(
            {
                "position_id": position_id,
                "symbol": position.get("symbol", ""),
                "entry_type": position.get("entry_type", ""),
                "timeframe": position.get("timeframe", ""),
                "strategy": position.get("actual_strategy", position.get("strategy", "")),
                "direction": "BUY" if entry is not None and int(entry.get("type", 0)) == 0 else "SELL",
                "mt5_entry_comment": entry.get("comment", "") if entry is not None else position.get("comment", ""),
                "opened_at_wib": _format_time(entry.get("time") if entry is not None else position.get("time_open")),
                "closed_at_wib": _format_time(exit_deal.get("time") if exit_deal is not None else None),
                "entry_price": entry.get("price") if entry is not None else None,
                "planned_entry": signal.get("planned_entry"),
                "planned_sl": signal.get("planned_sl"),
                "planned_tp": signal.get("planned_tp"),
                "exit_price": exit_deal.get("price") if exit_deal is not None else None,
                "mt5_exit_comment": exit_deal.get("comment", "") if exit_deal is not None else "",
                "net_profit_usd": net_profit,
                "outcome": "WIN" if is_win else "LOSS",
                "signal_probability": signal.get("signal_probability"),
                "planned_levels_status": "matched signal" if signal else "unmatched signal",
            }
        )

    columns = [
        "position_id", "symbol", "entry_type", "timeframe", "strategy", "direction", "mt5_entry_comment",
        "opened_at_wib", "closed_at_wib", "entry_price", "planned_entry", "planned_sl",
        "planned_tp", "exit_price", "mt5_exit_comment", "net_profit_usd", "outcome",
        "signal_probability", "planned_levels_status",
    ]
    return pd.DataFrame(records, columns=columns).sort_values("opened_at_wib", kind="mergesort")


def build_forward_summary(forward_trades: pd.DataFrame) -> pd.DataFrame:
    if forward_trades.empty:
        return pd.DataFrame(columns=["entry_type", "timeframe", "strategy", "trades", "wins", "losses", "winrate_pct", "net_profit_usd"])
    summary = forward_trades.groupby(["entry_type", "timeframe", "strategy"], dropna=False).agg(
        trades=("outcome", "size"),
        wins=("outcome", lambda values: int((values == "WIN").sum())),
        losses=("outcome", lambda values: int((values == "LOSS").sum())),
        winrate_pct=("outcome", lambda values: round(float((values == "WIN").mean() * 100), 2)),
        net_profit_usd=("net_profit_usd", "sum"),
    ).reset_index()
    return summary.sort_values(["entry_type", "timeframe", "strategy"], kind="mergesort")


def load_telegram_delivery_events(path: Path) -> pd.DataFrame:
    columns = ["sent_at_wib", "channel", "delivered", "message"]
    if not path.exists():
        return pd.DataFrame(columns=columns)
    records = []
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            continue
        records.append({column: record.get(column, "") for column in columns})
    return pd.DataFrame(records, columns=columns)


def _write_frame(worksheet, frame: pd.DataFrame) -> None:
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append(list(row))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_workbook(path: Path, standard_limit: pd.DataFrame, forward_trades: pd.DataFrame, forward_summary: pd.DataFrame, telegram_events: pd.DataFrame, generated_at: str) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Forex SMC Analyzer Public Evidence"])
    overview.append(["Generated", generated_at])
    overview.append(["Standard-limit evidence", "MT5 bid/ask real-tick replay; coverage is stored per row."])
    overview.append(["Forward evidence", "Closed MT5 trades include raw MT5 entry and exit comments. Planned SL/TP is blank when no source signal can be matched."])
    overview.append(["Telegram evidence", "Delivery journal is populated only after TELEGRAM_EVENT_LOG_ENABLED is enabled locally."])
    overview.append(["Warning", "Results are historical evidence, not a profit guarantee or trading recommendation."])
    overview.column_dimensions["A"].width = 30
    overview.column_dimensions["B"].width = 105
    overview["A1"].font = Font(size=14, bold=True)
    _write_frame(workbook.create_sheet("Standard Limit Real Tick"), standard_limit)
    _write_frame(workbook.create_sheet("Forward Trades"), forward_trades)
    _write_frame(workbook.create_sheet("Forward Summary"), forward_summary)
    _write_frame(workbook.create_sheet("Telegram Delivery"), telegram_events)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export public forward-test evidence and Excel workbook.")
    parser.add_argument("--positions", type=Path, default=DEFAULT_POSITIONS)
    parser.add_argument("--deals", type=Path, default=DEFAULT_DEALS)
    parser.add_argument("--standard-limit", type=Path, default=DEFAULT_STANDARD_LIMIT)
    parser.add_argument("--telegram-events", type=Path, default=DEFAULT_TELEGRAM_EVENTS)
    parser.add_argument("--output-dir", type=Path, default=REPORTS_DIR)
    args = parser.parse_args()

    if not args.positions.exists() or not args.deals.exists():
        raise SystemExit("Forward-test source files are missing. Generate the local MT5 trade analysis first.")
    if not args.standard_limit.exists():
        raise SystemExit("Real-tick standard-limit result is missing. Run src.real_tick_backtester first.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(ZoneInfo("Asia/Jakarta")).strftime("%Y-%m-%d %H:%M:%S WIB")
    standard_limit = pd.read_csv(args.standard_limit)
    forward_trades = build_forward_trades(args.positions, args.deals)
    forward_summary = build_forward_summary(forward_trades)
    telegram_events = load_telegram_delivery_events(args.telegram_events)

    standard_limit.to_csv(args.output_dir / "standard_limit_real_tick_may2026.csv", index=False)
    forward_trades.to_csv(args.output_dir / "forward_test_trades.csv", index=False)
    forward_summary.to_csv(args.output_dir / "forward_test_summary.csv", index=False)
    telegram_events.to_csv(args.output_dir / "telegram_delivery_events.csv", index=False)
    write_workbook(args.output_dir / "forward_test_report.xlsx", standard_limit, forward_trades, forward_summary, telegram_events, generated_at)

    metadata = {
        "generated_at": generated_at,
        "standard_limit_rows": int(len(standard_limit)),
        "forward_closed_trades": int(len(forward_trades)),
        "forward_signal_level_matches": int((forward_trades["planned_levels_status"] == "matched signal").sum()),
        "telegram_delivery_events": int(len(telegram_events)),
        "sources": {
            "standard_limit": str(args.standard_limit.relative_to(BASE_DIR)),
            "positions": str(args.positions.relative_to(BASE_DIR)),
            "deals": str(args.deals.relative_to(BASE_DIR)),
        },
    }
    (args.output_dir / "report_metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    print(json.dumps(metadata, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
